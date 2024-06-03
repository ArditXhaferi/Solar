import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from datetime import datetime

def create_spreadsheet(data, file_name='output.xlsx', sheet_name='Sheet1'):
    """
    Create a spreadsheet from a Python list, style the headers, add borders to the table, 
    align text, adjust column width based on content, and save it to a file.

    Parameters:
    - data: The Python list containing the data.
    - file_name: The name of the output spreadsheet file (default: 'output.xlsx').
    - sheet_name: The name of the sheet in the spreadsheet (default: 'Sheet1').
    """
    # Create a DataFrame from the list
    df = pd.DataFrame(data)
    
    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        # Write the DataFrame to the writer
        df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False, header=True)
        
        # Access the openpyxl workbook and sheet objects
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Add heading and current date
        heading = "Data:"
        current_date = datetime.now().strftime("%d %B %Y")
        worksheet['A1'] = heading
        worksheet['A2'] = current_date

        
        # Style configurations
        header_font = Font(bold=True, color="c00000", name="Arial")
        cell_font = Font(name="Arial")  # Default font for cells
        cell_main = Font(bold=True, name="Arial")  # Default font for cells
        header_fill = PatternFill("solid", fgColor="f2dcdb")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        left_align = Alignment(horizontal="left")
        right_align = Alignment(horizontal="right")
        worksheet['A1'].font = Font(bold=True, color="c00000", name="Arial")
        worksheet['A2'].font = Font(bold=True, color="c00000", name="Arial")
        worksheet['A1'].fill = header_fill
        worksheet['A2'].fill = header_fill
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet['A2'].alignment = Alignment(horizontal='center')
        # Apply styles and alignments
        for cell in worksheet[4]:  # Headers
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
                if cell.column == 1:
                    cell.font = cell_main
                else:
                    cell.font = cell_font  

                # Left-align the first two columns, right-align the rest
                if cell.column <= 2:
                    cell.alignment = left_align
                else:
                    cell.alignment = right_align

        # Adjust column width
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells) + 1
            worksheet.column_dimensions[column_cells[0].column_letter].width = length

    print(f"Spreadsheet created successfully: {file_name}, Sheet: {sheet_name}")
